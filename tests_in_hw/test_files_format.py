import zipfile
from zipfile import ZipFile
import csv
from pypdf import PdfReader
from openpyxl import load_workbook
from script_os import ZIP_PATH


def test_pdf():
    with ZipFile(ZIP_PATH, 'r') as zip_file:
        with zip_file.open("Адреса.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()
            assert "Металлургов" in text
            assert "321" in text
            assert "Санкт-Петербург" not in text

def test_csv():
    with ZipFile(ZIP_PATH, 'r') as zip_file:
        with zip_file.open("people-100.csv", "r") as csv_file:
            my_reader = csv_file.read().decode('utf-8-sig')
            assert "elijah57@example.net" in my_reader

def test_xlsx():
    with ZipFile(ZIP_PATH, 'r') as zip_file:
        with zip_file.open("Адреса.xlsx") as xlsx_file:
            workbook = load_workbook(filename=xlsx_file)
            ws = workbook.active
            values = ws.cell(row=6, column=2).value
            assert "Москва" in values
